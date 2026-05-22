import io
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from apps.account.permissions import MustChangePasswordPermission
from apps.account.utils import get_site_utilisateur
from apps.visites_periodiques.models import LigneVisitePeriodique, ListeVisitePeriodique
from django.utils import timezone


class ExportVisitesPeriodiquesView(APIView):
    permission_classes = [MustChangePasswordPermission, IsAuthenticated]

    def get(self, request):
        site = get_site_utilisateur(request.user)
        qs = LigneVisitePeriodique.objects.select_related(
            "liste", "liste__medecin__profile__user", "collaborateur"
        ).all()
        if site:
            qs = qs.filter(liste__medecin__site=site)

        today = timezone.localdate()
        wb = Workbook()
        ws = wb.active
        ws.title = "Visites périodiques"

        headers = [
            "Référence liste", "Date visite", "Collaborateur", "Matricule", "Présence", "Statut liste"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")

        for lv in qs.order_by("liste__date_visite", "collaborateur__matricule"):
            liste = lv.liste
            coll = lv.collaborateur
            ws.append([
                liste.reference,
                liste.date_visite.strftime("%d/%m/%Y") if liste.date_visite else "",
                f"{getattr(coll, 'nom', '')} {getattr(coll, 'prenom', '')}".strip(),
                getattr(coll, 'matricule', ''),
                lv.presence,
                liste.statut,
            ])

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return FileResponse(
            stream, as_attachment=True,
            filename=f"visites_periodiques_{today}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )