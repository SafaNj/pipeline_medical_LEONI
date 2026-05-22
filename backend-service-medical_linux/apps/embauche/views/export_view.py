import io

from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def export_liste_embauche(liste):
    wb = Workbook()
    ws = wb.active
    ws.title = liste.date_visite.strftime('%Y-%m-%d') if liste.date_visite else 'Liste Embauche'

    headers = [
        'N°',
        'Collaborateur',
        'Niveau',
        'Matricule',
        'CIN',
        'Date naissance',
        'num demande',
        'PS',
        'Projet',
        'Date recruttment',
        'centre de cout',
        'Fonction',
        "Source d'information",
        'Gouvernorat',
        'Genre',
        'Telephone',
        'Formation',
        'Présence',
        'Etat Embauche',
        'Observations médecin',
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='1F4E78')

    for index, candidat in enumerate(liste.candidats.all().order_by('id'), start=1):
        ws.append([
            index,
            f'{candidat.nom} {candidat.prenom}'.strip(),
            candidat.niveau,
            candidat.matricule,
            candidat.cin,
            candidat.date_naissance.strftime('%Y-%m-%d') if candidat.date_naissance else '',
            candidat.num_demande,
            candidat.ps,
            candidat.projet,
            candidat.date_recrutement.strftime('%Y-%m-%d') if candidat.date_recrutement else '',
            candidat.centre_cout,
            candidat.poste,
            candidat.source_information,
            candidat.gouvernorat,
            candidat.get_genre_display() if candidat.genre else '',
            candidat.telephone,
            candidat.formation,
            candidat.get_presence_display(),
            candidat.get_etat_embauche_display(),
            candidat.observations_medecin,
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f'liste_embauche_{liste.reference}.xlsx'
    return FileResponse(
        stream,
        as_attachment=True,
        filename=filename,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
