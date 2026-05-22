import io
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from apps.account.permissions import MustChangePasswordPermission
from apps.account.utils import get_site_utilisateur
from apps.act_infirmier.permissions import IsInfirmier
from apps.stock.models import StockMedicament
from django.utils import timezone


class ExportStockMedicamentsView(APIView):
    permission_classes = [MustChangePasswordPermission, IsAuthenticated, IsInfirmier]

    def get(self, request):
        site = get_site_utilisateur(request.user)
        qs = StockMedicament.objects.select_related("medicament", "site")
        if site:
            qs = qs.filter(site=site)

        today = timezone.localdate()
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventaire médicaments"

        headers = [
            "Médicament", "Quantité", "Seuil alerte",
            "Date expiration", "Statut", "Site"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")

        for sm in qs.order_by("medicament__nom"):
            exp = sm.date_expiration
            if sm.quantite == 0:
                statut = "Rupture"
            elif sm.quantite <= sm.seuil_alerte:
                statut = "Stock limite"
            elif exp and exp < today:
                statut = "Périmé"
            elif exp and (exp - today).days <= 90:
                statut = "Proche expiration"
            else:
                statut = "OK"

            ws.append([
                str(sm.medicament),
                sm.quantite,
                sm.seuil_alerte,
                exp.strftime("%d/%m/%Y") if exp else "",
                statut,
                str(sm.site) if sm.site else "",
            ])

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return FileResponse(
            stream, as_attachment=True,
            filename=f"inventaire_medicaments_{today}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
