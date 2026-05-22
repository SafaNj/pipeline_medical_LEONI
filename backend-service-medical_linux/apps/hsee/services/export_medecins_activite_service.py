"""
Export Excel — activité médecins (consultations, fiches aptitude, contre-visites).
Sans données cliniques sensibles (pas de diagnostic / notes).
"""
from __future__ import annotations

import logging
import math
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from typing import Any, Callable

from django.db.models import Count, Max
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.consultations.models import Consultation
from apps.control_visits.models import ContreVisite
from apps.account.utils import filter_queryset_by_site
from apps.medical_work.models import FicheAptitude
from apps.hsee.services.medecin_export_role import (
    TYPE_CONTROLEUR,
    TYPE_TRAITANT,
    TYPE_TRAVAIL,
)

logger = logging.getLogger(__name__)

MAX_PERIOD_JOURS = 366
MAX_LIGNES_DETAIL = 50_000


def _site_scope(queryset, site):
    return filter_queryset_by_site(queryset, site)

def _medecin_nom_affiche(medecin) -> str:
    if not medecin:
        return "—"
    user = medecin.profile.user
    nom = (user.get_full_name() or "").strip()
    if not nom:
        nom = (user.username or "").strip()
    if not nom:
        return "—"
    if nom.lower().startswith("dr."):
        return nom
    return f"Dr. {nom}"


def _normalise_matricule(s: str | None) -> str:
    return (s or "").strip().upper()


def _cell(v: Any) -> Any:
    """Valeur sûre pour Excel (pas de NaN ; None → chaîne vide)."""
    if v is None:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    if isinstance(v, bool):
        return "oui" if v else "non"
    return v


def _collab_rh_fields(collab) -> dict[str, Any]:
    if not collab:
        return {
            "collab_poste": "",
            "collab_departement": "",
            "collab_telephone": "",
            "collab_email": "",
            "collab_date_naissance": "",
            "collab_anciennete": "",
        }
    poste, dept, tel, email = "", "", "", ""
    try:
        poste = collab.poste or ""
    except Exception:
        pass
    try:
        dept = collab.department or ""
    except Exception:
        pass
    try:
        tel = collab.telephone or ""
    except Exception:
        pass
    try:
        email = collab.email or ""
    except Exception:
        pass
    dn = ""
    try:
        if collab.date_naissance:
            dn = collab.date_naissance.isoformat()
    except Exception:
        pass
    ancien = ""
    try:
        if collab.date_embauche:
            delta = date.today() - collab.date_embauche
            ancien = f"{delta.days // 365} an(s)" if delta.days >= 365 else f"{delta.days} j."
    except Exception:
        pass
    return {
        "collab_poste": poste,
        "collab_departement": dept,
        "collab_telephone": tel,
        "collab_email": email,
        "collab_date_naissance": dn,
        "collab_anciennete": ancien,
    }


def _oui_non_predicat(pred: bool) -> str:
    return "oui" if pred else "non"


def _cle_patient(collab_id: int | None, matricule: str | None, modele: str, pk: int) -> str:
    if collab_id:
        return f"COLLAB:{collab_id}"
    m = _normalise_matricule(matricule)
    if m:
        return f"MAT:{m}"
    return f"INCONN:{modele.upper()}:{pk}"


def _parse_type_medecin_param(val: str | None) -> set[str] | None:
    if not val or not val.strip():
        return None
    v = val.strip().lower()
    if v == TYPE_TRAITANT:
        return {TYPE_TRAITANT}
    if v == TYPE_TRAVAIL:
        return {TYPE_TRAVAIL}
    if v == TYPE_CONTROLEUR:
        return {TYPE_CONTROLEUR}
    return None  # invalide → traiter comme « tous »


def collecter_lignes_detail(
    date_debut: date,
    date_fin: date,
    types_medecin: set[str] | None,
    medecin_id: int | None,
    site=None,
) -> tuple[list[dict[str, Any]], str | None]:
    """
    Retourne (lignes, erreur). Si erreur non vide, ne pas générer le fichier.
    """
    lignes: list[dict[str, Any]] = []

    inclure_traitant = types_medecin is None or TYPE_TRAITANT in types_medecin
    inclure_travail = types_medecin is None or TYPE_TRAVAIL in types_medecin
    inclure_controleur = types_medecin is None or TYPE_CONTROLEUR in types_medecin

    if inclure_traitant:
        qs = _site_scope(
            Consultation.objects.filter(
                date_consultation__date__gte=date_debut,
                date_consultation__date__lte=date_fin,
            )
            .select_related("medecin", "medecin__med_type", "item_passage__collaborateur")
            .annotate(
                _nb_ordonnances=Count("ordonnances", distinct=True),
                _nb_certificats=Count("certificats", distinct=True),
                _max_jours_certificat=Max("certificats__jours_repos"),
                _max_date_debut_repos=Max("certificats__date_debut_repos"),
            )
            .order_by("date_consultation"),
            site,
        )
        if medecin_id:
            qs = qs.filter(medecin_id=medecin_id)
        for c in qs.iterator(chunk_size=500):
            if len(lignes) >= MAX_LIGNES_DETAIL:
                return [], f"Limite dépassée : maximum {MAX_LIGNES_DETAIL} lignes de détail pour la période."
            collab = c.item_passage.collaborateur if c.item_passage_id else None
            cid = collab.id if collab else None
            mat = collab.matricule if collab else None
            nom_prenom = ""
            if collab:
                try:
                    nom_prenom = f"{collab.nom or ''} {collab.prenom or ''}".strip()
                except Exception:
                    nom_prenom = ""
            src = "ITEM_PASSAGE" if collab else "PASSAGE_SANS_COLLABORATEUR"
            extra = _collab_rh_fields(collab)
            nb_ord = getattr(c, "_nb_ordonnances", 0) or 0
            nb_cert = getattr(c, "_nb_certificats", 0) or 0
            max_j = getattr(c, "_max_jours_certificat", None)
            max_date_repos = getattr(c, "_max_date_debut_repos", None)
            lignes.append(
                {
                    "date_acte": c.date_consultation,
                    "date_jour": c.date_consultation.date(),
                    "type_acte": "CONSULTATION",
                    "type_medecin_role": TYPE_TRAITANT,
                    "medecin_id": c.medecin_id,
                    "medecin_nom": _medecin_nom_affiche(c.medecin),
                    "collaborateur_id": cid,
                    "matricule": mat or "",
                    "nom_prenom": nom_prenom or "",
                    "source_identification": src,
                    "type_visite": "",
                    "type_visite_libelle": "",
                    "duree_repos": "",
                    "refus_repos": "",
                    "ordonnance_oui": _oui_non_predicat(nb_ord > 0),
                    "certificat_oui": _oui_non_predicat(nb_cert > 0),
                    "certificat_jours_arret": max_j if max_j is not None else "",
                    "certificat_date_debut_repos": (
                        max_date_repos.isoformat() if max_date_repos else ""
                    ),
                    "cv_a_partir": "",
                    "cv_repos_initial": "",
                    "cv_ecart_repos": "",
                    "cv_remarque": "",
                    "aptitude": "",
                    "precision_aptitude": "",
                    "lien_medecin_traitant_nom": "",
                    **extra,
                    "cle_patient": _cle_patient(cid, mat, "consultation", c.pk),
                    "pk": c.pk,
                }
            )

    if inclure_travail:
        qs = _site_scope(
            FicheAptitude.objects.filter(date_visite__gte=date_debut, date_visite__lte=date_fin)
            .select_related("medecin_travail", "medecin_travail__med_type", "collaborateur")
            .order_by("date_visite", "id"),
            site,
        )
        if medecin_id:
            qs = qs.filter(medecin_travail_id=medecin_id)
        for f in qs.iterator(chunk_size=500):
            if len(lignes) >= MAX_LIGNES_DETAIL:
                return [], f"Limite dépassée : maximum {MAX_LIGNES_DETAIL} lignes de détail pour la période."
            cid = f.collaborateur_id
            if cid:
                collab = f.collaborateur
                mat = collab.matricule if collab else None
                try:
                    nom_prenom = f"{collab.nom or ''} {collab.prenom or ''}".strip()
                except Exception:
                    nom_prenom = ""
                src = "FICHE_COLLABORATEUR"
            else:
                mat = f.matricule or ""
                nom_prenom = ""
                src = "EMBAUCHE_SANS_COLLABORATEUR"
            extra = _collab_rh_fields(collab if cid else None)
            try:
                tv_label = f.get_type_visite_display()
            except Exception:
                tv_label = f.type_visite or ""
            try:
                aptitude_label = f.get_aptitude_display()
            except Exception:
                aptitude_label = f.aptitude or ""
            lignes.append(
                {
                    "date_acte": timezone.make_aware(
                        datetime.combine(f.date_visite, datetime.min.time())
                    ),
                    "date_jour": f.date_visite,
                    "type_acte": "FICHE_APTITUDE",
                    "type_medecin_role": TYPE_TRAVAIL,
                    "medecin_id": f.medecin_travail_id,
                    "medecin_nom": _medecin_nom_affiche(f.medecin_travail),
                    "collaborateur_id": cid,
                    "matricule": mat or "",
                    "nom_prenom": nom_prenom,
                    "source_identification": src,
                    "type_visite": f.type_visite,
                    "type_visite_libelle": tv_label,
                    "duree_repos": "",
                    "refus_repos": "",
                    "ordonnance_oui": "",
                    "certificat_oui": "",
                    "certificat_jours_arret": "",
                    "certificat_date_debut_repos": "",
                    "cv_a_partir": "",
                    "cv_repos_initial": "",
                    "cv_ecart_repos": "",
                    "cv_remarque": "",
                    "aptitude": aptitude_label,
                    "precision_aptitude": f.precision_aptitude or "",
                    "lien_medecin_traitant_nom": "",
                    **extra,
                    "cle_patient": _cle_patient(cid, mat or f.matricule, "fiche", f.pk),
                    "pk": f.pk,
                }
            )

    if inclure_controleur:
        qs = _site_scope(
            ContreVisite.objects.filter(date__gte=date_debut, date__lte=date_fin)
            .select_related(
                "medecin_controleur",
                "medecin_controleur__med_type",
                "item_passage__collaborateur",
                "item_passage__liste__medecin",
                "item_passage__liste__medecin__profile__user",
            )
            .order_by("date", "id"),
            site,
        )
        if medecin_id:
            qs = qs.filter(medecin_controleur_id=medecin_id)
        for cv in qs.iterator(chunk_size=500):
            if len(lignes) >= MAX_LIGNES_DETAIL:
                return [], f"Limite dépassée : maximum {MAX_LIGNES_DETAIL} lignes de détail pour la période."
            collab = (
                cv.item_passage.collaborateur
                if cv.item_passage_id and cv.item_passage.collaborateur_id
                else None
            )
            if collab:
                cid = collab.id
                mat = collab.matricule
                try:
                    nom_prenom = f"{collab.nom or ''} {collab.prenom or ''}".strip()
                except Exception:
                    nom_prenom = ""
                src = "ITEM_PASSAGE"
            else:
                cid = None
                mat = _normalise_matricule(cv.matricule) or cv.matricule
                nom_prenom = cv.nom_prenom or ""
                src = "MATRICULE_BRUT"
            extra = _collab_rh_fields(collab if collab else None)
            mt_nom = ""
            if cv.item_passage_id and cv.item_passage.liste_id and cv.item_passage.liste.medecin_id:
                mt_nom = _medecin_nom_affiche(cv.item_passage.liste.medecin)
            ecart = ""
            if cv.repos_initial is not None:
                try:
                    ecart = str(int(cv.duree_repos) - int(cv.repos_initial))
                except (TypeError, ValueError):
                    ecart = ""
            lignes.append(
                {
                    "date_acte": timezone.make_aware(
                        datetime.combine(cv.date, datetime.min.time())
                    ),
                    "date_jour": cv.date,
                    "type_acte": "CONTRE_VISITE",
                    "type_medecin_role": TYPE_CONTROLEUR,
                    "medecin_id": cv.medecin_controleur_id,
                    "medecin_nom": _medecin_nom_affiche(cv.medecin_controleur),
                    "collaborateur_id": cid,
                    "matricule": mat or "",
                    "nom_prenom": nom_prenom,
                    "source_identification": src,
                    "type_visite": "",
                    "type_visite_libelle": "",
                    "duree_repos": cv.duree_repos,
                    "refus_repos": cv.refus_repos,
                    "ordonnance_oui": "",
                    "certificat_oui": "",
                    "certificat_jours_arret": "",
                    "certificat_date_debut_repos": "",
                    "cv_a_partir": cv.a_partir.isoformat() if cv.a_partir else "",
                    "cv_repos_initial": cv.repos_initial if cv.repos_initial is not None else "",
                    "cv_ecart_repos": ecart,
                    "cv_remarque": cv.remarque or "",
                    "aptitude": "",
                    "precision_aptitude": "",
                    "lien_medecin_controlleur_nom": mt_nom,
                    **extra,
                    "cle_patient": _cle_patient(cid, mat, "contre_visite", cv.pk),
                    "pk": cv.pk,
                }
            )

    return lignes, None


def aggregats(lignes: list[dict[str, Any]]) -> tuple[list[dict], list[dict]]:
    """Synthèse par (role, medecin_id) et par mois (yyyy-mm, role, medecin_id)."""
    synthese: dict[tuple[str, int | None], dict[str, Any]] = defaultdict(
        lambda: {"nb_actes": 0, "cles": set(), "jours": set(), "medecin_nom": ""}
    )
    par_mois: dict[tuple[str, str, int | None], dict[str, Any]] = defaultdict(
        lambda: {"nb_actes": 0, "cles": set(), "medecin_nom": ""}
    )

    for row in lignes:
        role = row["type_medecin_role"]
        mid = row["medecin_id"]
        k = (role, mid)
        synthese[k]["nb_actes"] += 1
        synthese[k]["cles"].add(row["cle_patient"])
        synthese[k]["jours"].add(row["date_jour"])
        synthese[k]["medecin_nom"] = row.get("medecin_nom") or synthese[k]["medecin_nom"]

        mois = row["date_jour"].strftime("%Y-%m")
        km = (mois, role, mid)
        par_mois[km]["nb_actes"] += 1
        par_mois[km]["cles"].add(row["cle_patient"])
        par_mois[km]["medecin_nom"] = row.get("medecin_nom") or par_mois[km]["medecin_nom"]

    synthese_rows = []
    for (role, mid), v in sorted(synthese.items(), key=lambda x: (x[0][0], x[0][1] or -1)):
        nb_jours = len(v["jours"])
        nb_actes = v["nb_actes"]
        synthese_rows.append(
            {
                "type_medecin": role,
                "medecin_nom": v.get("medecin_nom") or "—",
                "nb_actes_total": nb_actes,
                "nb_collaborateurs_uniques": len(v["cles"]),
                "nb_jours_avec_acte": nb_jours,
                "actes_par_jour_actif": round(nb_actes / nb_jours, 4) if nb_jours else 0,
            }
        )

    mois_rows = []
    for (mois, role, mid), v in sorted(par_mois.items(), key=lambda x: (x[0][0], x[0][1], x[0][2] or -1)):
        mois_rows.append(
            {
                "mois": mois,
                "type_medecin": role,
                "medecin_nom": v.get("medecin_nom") or "—",
                "nb_actes": v["nb_actes"],
                "nb_collaborateurs_uniques": len(v["cles"]),
            }
        )

    return synthese_rows, mois_rows


def _detail_columns_for_role(
    role: str,
) -> list[tuple[str, Callable[[dict[str, Any]], Any]]]:
    if role == TYPE_TRAITANT:
        return [
            (
                "Date et heure de consultation",
                lambda row: row["date_acte"].isoformat()
                if hasattr(row["date_acte"], "isoformat")
                else row["date_acte"],
            ),
            (
                "Date de consultation",
                lambda row: row["date_jour"].isoformat()
                if hasattr(row["date_jour"], "isoformat")
                else row["date_jour"],
            ),
            ("Médecin traitant", lambda row: row.get("medecin_nom", "")),
            ("Matricule", lambda row: row.get("matricule", "")),
            ("Nom et prénom", lambda row: row.get("nom_prenom", "")),
            ("Poste", lambda row: row.get("collab_poste", "")),
            ("Département", lambda row: row.get("collab_departement", "")),
            ("Date de naissance", lambda row: row.get("collab_date_naissance", "")),
            ("Ancienneté", lambda row: row.get("collab_anciennete", "")),
            ("Ordonnance émise", lambda row: row.get("ordonnance_oui", "")),
            ("Certificat émis", lambda row: row.get("certificat_oui", "")),
            (
                "Date début repos (certificat traitant)",
                lambda row: row.get("certificat_date_debut_repos", ""),
            ),
            (
                "Jours d'arrêt (certificat)",
                lambda row: row.get("certificat_jours_arret", ""),
            ),
        ]

    if role == TYPE_TRAVAIL:
        return [
            (
                "Date de visite",
                lambda row: row["date_jour"].isoformat()
                if hasattr(row["date_jour"], "isoformat")
                else row["date_jour"],
            ),
            ("Médecin du travail", lambda row: row.get("medecin_nom", "")),
            ("Matricule", lambda row: row.get("matricule", "")),
            ("Nom et prénom", lambda row: row.get("nom_prenom", "")),
            ("Poste", lambda row: row.get("collab_poste", "")),
            ("Département", lambda row: row.get("collab_departement", "")),
            ("Date de naissance", lambda row: row.get("collab_date_naissance", "")),
            ("Ancienneté", lambda row: row.get("collab_anciennete", "")),
            ("Type de visite", lambda row: row.get("type_visite_libelle", "")),
            ("Aptitude", lambda row: row.get("aptitude", "")),
            ("Précision aptitude", lambda row: row.get("precision_aptitude", "")),
        ]

    if role == TYPE_CONTROLEUR:
        return [
            (
                "Date de contre-visite",
                lambda row: row["date_jour"].isoformat()
                if hasattr(row["date_jour"], "isoformat")
                else row["date_jour"],
            ),
            ("Médecin contrôleur", lambda row: row.get("medecin_nom", "")),
            ("Matricule", lambda row: row.get("matricule", "")),
            ("Nom et prénom", lambda row: row.get("nom_prenom", "")),
            ("Poste", lambda row: row.get("collab_poste", "")),
            ("Département", lambda row: row.get("collab_departement", "")),
            ("Date début repos (contre-visite)", lambda row: row.get("cv_a_partir", "")),
            (
                "Repos initial ",
                lambda row: row.get("cv_repos_initial", ""),
            ),
            (
                "Écart en jours vs repos initial",
                lambda row: row.get("cv_ecart_repos", ""),
            ),
            ("Refus de repos", lambda row: row.get("refus_repos", "")),
            ("Durée de repos finale", lambda row: row.get("duree_repos", "")),
            ("Remarque du contrôleur", lambda row: row.get("cv_remarque", "")),
            ("Médecin contrôleur lié", lambda row: row.get("lien_medecin_controlleur houwa hh" \
            "_nom", "")),
        ]

    raise ValueError(f"Role medecin inconnu: {role}")


def _detail_sheet_name(role: str) -> str:
    if role == TYPE_TRAITANT:
        return "Traitant - Consultations"
    if role == TYPE_TRAVAIL:
        return "Travail - Fiches aptitude"
    if role == TYPE_CONTROLEUR:
        return "Contrôleur - Contre-visites"
    return "Détail"


def _detail_colors(role: str) -> tuple[str, str]:
    if role == TYPE_TRAITANT:
        return "2E7D32", "F1F8E9"
    if role == TYPE_TRAVAIL:
        return "1565C0", "E3F2FD"
    if role == TYPE_CONTROLEUR:
        return "E65100", "FFF3E0"
    return "37474F", "ECEFF1"


def _column_width_for_header(header: str) -> float:
    h = (header or "").strip().lower()
    if "date" in h or "mois" in h:
        return 22
    if (
        "nom" in h
        or "médecin" in h
        or "medecin" in h
        or "département" in h
        or "departement" in h
        or "précision" in h
        or "precision" in h
        or "remarque" in h
    ):
        return 28
    if (
        h.startswith("nb ")
        or "total" in h
        or "moyenne" in h
        or "jours" in h
        or "durée" in h
        or "duree" in h
        or "écart" in h
        or "ecart" in h
    ):
        return 14
    return 18


def _appliquer_mise_en_forme(ws, couleur_header: str, couleur_zebre: str) -> None:
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    fill_header = PatternFill(fill_type="solid", fgColor=couleur_header)
    fill_zebre = PatternFill(fill_type="solid", fgColor=couleur_zebre)
    font_header = Font(color="FFFFFF", bold=True)
    align_cell = Alignment(wrap_text=True, vertical="center")
    border_thin = Border(
        left=Side(style="thin", color="CFD8DC"),
        right=Side(style="thin", color="CFD8DC"),
        top=Side(style="thin", color="CFD8DC"),
        bottom=Side(style="thin", color="CFD8DC"),
    )

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_cell
        cell.border = border_thin
        header = str(cell.value or "")
        ws.column_dimensions[get_column_letter(col)].width = _column_width_for_header(header)

    for row_idx in range(2, ws.max_row + 1):
        zebra = row_idx % 2 == 0
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = align_cell
            cell.border = border_thin
            if zebra:
                cell.fill = fill_zebre


def _write_detail_sheet(ws, lignes: list[dict[str, Any]], role: str) -> None:
    det_columns = _detail_columns_for_role(role)
    ws.append([header for header, _ in det_columns])
    for row in lignes:
        if row.get("type_medecin_role") != role:
            continue
        ws.append([_cell(getter(row)) for _, getter in det_columns])


def build_medecins_activite_workbook_bytes(
    date_debut: date,
    date_fin: date,
    type_medecin: str | None,
    medecin_id: int | None,
    site=None,
) -> tuple[bytes | None, str | None, int | None]:
    """
    Retourne (contenu xlsx ou None, message_erreur ou None, nb_lignes_detail ou None).
    """
    types = _parse_type_medecin_param(type_medecin)
    lignes, err = collecter_lignes_detail(date_debut, date_fin, types, medecin_id, site)
    if err:
        return None, err, None

    synthese_rows, mois_rows = aggregats(lignes)

    wb = Workbook()
    ws_syn = wb.active
    ws_syn.title = "Synthèse globale"
    ws_syn.append(
        [
            "Rôle du médecin",
            "Nom du médecin",
            "Total actes sur la période",
            "Nb collaborateurs suivis",
            "Nb jours de présence active",
            "Moyenne actes / jour travaillé",
        ]
    )
    for r in synthese_rows:
        ws_syn.append(
            [
                r["type_medecin"],
                r["medecin_nom"],
                r["nb_actes_total"],
                r["nb_collaborateurs_uniques"],
                r["nb_jours_avec_acte"],
                r["actes_par_jour_actif"],
            ]
        )
    _appliquer_mise_en_forme(ws_syn, "37474F", "ECEFF1")

    ws_mois = wb.create_sheet("Activité par mois")
    ws_mois.append(
        [
            "Mois (AAAA-MM)",
            "Rôle du médecin",
            "Nom du médecin",
            "Nb actes du mois",
            "Nb collaborateurs du mois",
        ]
    )
    for r in mois_rows:
        ws_mois.append(
            [
                r["mois"],
                r["type_medecin"],
                r["medecin_nom"],
                r["nb_actes"],
                r["nb_collaborateurs_uniques"],
            ]
        )
    _appliquer_mise_en_forme(ws_mois, "37474F", "ECEFF1")

    role_unique = next(iter(types)) if types and len(types) == 1 else None
    if role_unique:
        ws_det = wb.create_sheet(_detail_sheet_name(role_unique))
        _write_detail_sheet(ws_det, lignes, role_unique)
        couleur_header, couleur_zebre = _detail_colors(role_unique)
        _appliquer_mise_en_forme(ws_det, couleur_header, couleur_zebre)
    else:
        for role in (TYPE_TRAITANT, TYPE_TRAVAIL, TYPE_CONTROLEUR):
            ws_det = wb.create_sheet(_detail_sheet_name(role))
            _write_detail_sheet(ws_det, lignes, role)
            couleur_header, couleur_zebre = _detail_colors(role)
            _appliquer_mise_en_forme(ws_det, couleur_header, couleur_zebre)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), None, len(lignes)


def validate_export_params(
    date_debut: date | None, date_fin: date | None
) -> str | None:
    if not date_debut or not date_fin:
        return "Paramètres date_debut et date_fin requis (YYYY-MM-DD)."
    if date_debut > date_fin:
        return "date_debut doit être inférieur ou égal à date_fin."
    if (date_fin - date_debut).days > MAX_PERIOD_JOURS:
        return f"Période trop large : maximum {MAX_PERIOD_JOURS} jours."
    return None

