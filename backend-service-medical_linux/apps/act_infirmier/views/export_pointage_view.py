import io
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from apps.account.permissions import IsAnyMedecinOrHSSE, MustChangePasswordPermission
from apps.account.utils import get_site_utilisateur
from apps.act_infirmier.models import PointageMedecin
from django.utils import timezone


class ExportPointageMedecinView(APIView):
    permission_classes = [MustChangePasswordPermission, IsAuthenticated, IsAnyMedecinOrHSSE]

    def get(self, request):
        mois = request.query_params.get("mois")
        annee = request.query_params.get("annee")
        site = get_site_utilisateur(request.user)

        qs = PointageMedecin.objects.select_related("medecin__profile__user", "site")
        if site:
            qs = qs.filter(medecin__site=site)
        if mois:
            qs = qs.filter(mois=int(mois))
        if annee:
            qs = qs.filter(annee=int(annee))

        wb = Workbook()
        ws = wb.active
        ws.title = "Présence médecins"
        headers = ["Médecin", "Date", "Heures travaillées", "Remarque"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")

        for p in qs.order_by("medecin__profile__user__last_name", "date"):
            try:
                u = p.medecin.profile.user
                nom = f"Dr. {u.first_name} {u.last_name}".strip()
            except Exception:
                nom = f"Médecin #{p.medecin_id}"
            ws.append([nom, p.date.strftime("%d/%m/%Y"), p.heures_travaillees, p.remarque])

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        today = timezone.localdate()
        return FileResponse(
            stream, as_attachment=True,
            filename=f"presence_medecins_{today}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
